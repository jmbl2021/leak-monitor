"""Excel export functionality for leak-monitor.

Generates formatted XLSX spreadsheets with victim data.
Includes attribution to RansomLook.io per CC BY 4.0 license.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..config import get_config
from ..models import Victim, CompanyType, ReviewStatus

logger = logging.getLogger(__name__)

# Style definitions
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SEC_REGULATED_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
# 8-K disclosure timing fills
DISCLOSURE_COMPLIANT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green - <=4 days
DISCLOSURE_LATE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow - 5-14 days
DISCLOSURE_VERY_LATE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red - >14 days
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Column configuration
COLUMNS = [
    ("Post Date", 12),
    ("Group", 15),
    ("Victim (Raw)", 35),
    ("Company Name", 30),
    ("Type", 12),
    ("Region", 15),
    ("Country", 15),
    ("SEC Regulated", 12),
    ("CIK", 12),
    # 8-K columns
    ("8-K Filed", 10),
    ("8-K Date", 12),
    ("Disclosure Days", 14),
    # Remaining columns
    ("Subsidiary", 10),
    ("Parent Company", 25),
    ("ADR", 8),
    ("Status", 10),
    ("Notes", 50),
]


def create_victims_export(
    victims: list[Victim],
    filename: Optional[str] = None,
    title: Optional[str] = None
) -> Path:
    """Create an Excel export of victim data.

    Args:
        victims: List of Victim records to export
        filename: Optional custom filename (without extension)
        title: Optional title for the report

    Returns:
        Path to the generated Excel file
    """
    config = get_config()
    export_dir = Path(config.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename
    if not filename:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"victims_{timestamp}"

    filepath = export_dir / f"{filename}.xlsx"

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Victims"

    # Add title and metadata
    _add_header_section(ws, title, len(victims))

    # Add data table starting at row 5
    _add_data_table(ws, victims, start_row=5)

    # Add summary sheet
    _add_summary_sheet(wb, victims)

    # Add attribution sheet (CC BY 4.0 requirement)
    _add_attribution_sheet(wb)

    # Save workbook
    wb.save(filepath)
    logger.info(f"Exported {len(victims)} victims to {filepath}")

    return filepath


def _add_header_section(ws: Worksheet, title: Optional[str], count: int) -> None:
    """Add title and metadata to the worksheet."""
    # Title
    report_title = title or "Leak Monitor - Victim Report"
    ws["A1"] = report_title
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:N1")

    # Generated timestamp
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(size=10, italic=True)

    # Record count
    ws["A3"] = f"Total Records: {count}"
    ws["A3"].font = Font(size=10)

    # Blank row 4 for spacing


def _add_data_table(ws: Worksheet, victims: list[Victim], start_row: int) -> None:
    """Add the main data table to the worksheet."""
    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, victim in enumerate(victims, start=start_row + 1):
        # Format 8-K fields
        if victim.has_8k_filing is True:
            filed_8k = "Yes"
        elif victim.has_8k_filing is False:
            filed_8k = "No"
        else:
            filed_8k = "Unknown"

        sec_8k_date_str = victim.sec_8k_date.strftime("%Y-%m-%d") if victim.sec_8k_date else ""
        disclosure_days_str = str(victim.disclosure_days) if victim.disclosure_days is not None else ""

        row_data = [
            victim.post_date.strftime("%Y-%m-%d") if victim.post_date else "",
            victim.group_name,
            victim.victim_raw,
            victim.company_name or "",
            victim.company_type.value if victim.company_type else "",
            victim.region or "",
            victim.country or "",
            "Yes" if victim.is_sec_regulated else "No",
            victim.sec_cik or "",
            # 8-K columns
            filed_8k,
            sec_8k_date_str,
            disclosure_days_str,
            # Remaining columns
            "Yes" if victim.is_subsidiary else "No",
            victim.parent_company or "",
            "Yes" if victim.has_adr else "No",
            victim.review_status.value if victim.review_status else "",
            victim.notes or "",
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 17))  # Wrap notes

            # Conditional formatting for 8-K disclosure timing (column 12 = Disclosure Days)
            if col_idx == 12 and victim.disclosure_days is not None:
                if victim.disclosure_days <= 4:
                    cell.fill = DISCLOSURE_COMPLIANT_FILL
                elif victim.disclosure_days <= 14:
                    cell.fill = DISCLOSURE_LATE_FILL
                else:
                    cell.fill = DISCLOSURE_VERY_LATE_FILL
            # Standard conditional formatting
            elif victim.review_status == ReviewStatus.PENDING:
                cell.fill = PENDING_FILL
            elif victim.is_sec_regulated:
                cell.fill = SEC_REGULATED_FILL
            elif (row_idx - start_row) % 2 == 0:
                cell.fill = ALT_ROW_FILL

    # Freeze header row
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Enable auto-filter
    if victims:
        last_row = start_row + len(victims)
        last_col = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{last_row}"


def _add_summary_sheet(wb: Workbook, victims: list[Victim]) -> None:
    """Add a summary statistics sheet."""
    ws = wb.create_sheet(title="Summary")

    # Title
    ws["A1"] = "Summary Statistics"
    ws["A1"].font = Font(size=14, bold=True)

    row = 3

    # Total counts
    ws.cell(row=row, column=1, value="Total Victims:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=len(victims))
    row += 1

    pending = sum(1 for v in victims if v.review_status == ReviewStatus.PENDING)
    ws.cell(row=row, column=1, value="Pending Review:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=pending)
    row += 2

    # By company type
    ws.cell(row=row, column=1, value="By Company Type").font = Font(bold=True, underline="single")
    row += 1
    for ctype in CompanyType:
        count = sum(1 for v in victims if v.company_type == ctype)
        ws.cell(row=row, column=1, value=f"  {ctype.value.title()}:")
        ws.cell(row=row, column=2, value=count)
        row += 1
    row += 1

    # SEC regulated
    sec_count = sum(1 for v in victims if v.is_sec_regulated)
    ws.cell(row=row, column=1, value="SEC Regulated:").font = Font(bold=True)
    ws.cell(row=row, column=2, value=sec_count)
    row += 2

    # 8-K Statistics
    ws.cell(row=row, column=1, value="SEC 8-K Filings").font = Font(bold=True, underline="single")
    row += 1

    with_8k = sum(1 for v in victims if v.has_8k_filing is True)
    without_8k = sum(1 for v in victims if v.has_8k_filing is False)
    unknown_8k = sum(1 for v in victims if v.has_8k_filing is None and v.is_sec_regulated)

    ws.cell(row=row, column=1, value="  8-K Filed:")
    ws.cell(row=row, column=2, value=with_8k)
    row += 1
    ws.cell(row=row, column=1, value="  No 8-K Found:")
    ws.cell(row=row, column=2, value=without_8k)
    row += 1
    ws.cell(row=row, column=1, value="  Not Checked:")
    ws.cell(row=row, column=2, value=unknown_8k)
    row += 1

    # 8-K Disclosure timing breakdown
    if with_8k > 0:
        compliant = sum(1 for v in victims if v.disclosure_days is not None and v.disclosure_days <= 4)
        late = sum(1 for v in victims if v.disclosure_days is not None and 5 <= v.disclosure_days <= 14)
        very_late = sum(1 for v in victims if v.disclosure_days is not None and v.disclosure_days > 14)

        ws.cell(row=row, column=1, value="  <=4 days (compliant):")
        ws.cell(row=row, column=2, value=compliant)
        row += 1
        ws.cell(row=row, column=1, value="  5-14 days (late):")
        ws.cell(row=row, column=2, value=late)
        row += 1
        ws.cell(row=row, column=1, value="  >14 days (very late):")
        ws.cell(row=row, column=2, value=very_late)
        row += 1

    row += 1

    # Missing CIK warnings
    missing_cik = [v for v in victims if v.is_sec_regulated and not v.sec_cik]
    if missing_cik:
        ws.cell(row=row, column=1, value="Missing CIK Numbers").font = Font(bold=True, underline="single", color="FF0000")
        row += 1
        for v in missing_cik[:10]:
            ws.cell(row=row, column=1, value=f"  {v.company_name or v.victim_raw}")
            row += 1
        if len(missing_cik) > 10:
            ws.cell(row=row, column=1, value=f"  ... and {len(missing_cik) - 10} more")
            row += 1
        row += 1

    # By group (top 10)
    ws.cell(row=row, column=1, value="By Ransomware Group").font = Font(bold=True, underline="single")
    row += 1

    group_counts = {}
    for v in victims:
        group_counts[v.group_name] = group_counts.get(v.group_name, 0) + 1

    sorted_groups = sorted(group_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    for group, count in sorted_groups:
        ws.cell(row=row, column=1, value=f"  {group}:")
        ws.cell(row=row, column=2, value=count)
        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10


def _add_attribution_sheet(wb: Workbook) -> None:
    """Add attribution sheet per CC BY 4.0 license requirement."""
    ws = wb.create_sheet(title="Attribution")

    ws["A1"] = "Data Attribution"
    ws["A1"].font = Font(size=14, bold=True)

    ws["A3"] = "Data Source:"
    ws["B3"] = "RansomLook.io"
    ws["A3"].font = Font(bold=True)

    ws["A4"] = "Website:"
    ws["B4"] = "https://www.ransomlook.io"
    ws["A4"].font = Font(bold=True)

    ws["A5"] = "License:"
    ws["B5"] = "Creative Commons Attribution 4.0 (CC BY 4.0)"
    ws["A5"].font = Font(bold=True)

    ws["A6"] = "License URL:"
    ws["B6"] = "https://creativecommons.org/licenses/by/4.0/"
    ws["A6"].font = Font(bold=True)

    ws["A8"] = (
        "This data is sourced from RansomLook.io, which tracks ransomware "
        "group leak sites. The data is provided under the CC BY 4.0 license, "
        "which requires attribution when sharing or adapting the data."
    )
    ws.merge_cells("A8:D8")
    ws["A8"].alignment = Alignment(wrap_text=True)

    ws["A10"] = f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A10"].font = Font(italic=True)

    # Set column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 50
